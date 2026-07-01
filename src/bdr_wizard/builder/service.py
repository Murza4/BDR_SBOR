from __future__ import annotations

from collections import defaultdict
from pathlib import Path
from time import perf_counter
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
try:
    import pandera.pandas as pa
except ImportError:  # pragma: no cover - dependency is declared, fallback keeps local imports readable.
    pa = None

from bdr_wizard.excel.loader import WorkbookLoader
from bdr_wizard.models import ImportDecision, ImportSession, MappingRule, PerformanceEntry, Severity
from bdr_wizard.security import MAX_BUILD_ROWS, neutralize_excel_formula


class BdrBuilder:
    def __init__(
        self,
        output_root: Path = Path("data/outputs"),
        loader: WorkbookLoader | None = None,
    ) -> None:
        self.output_root = output_root
        self.loader = loader or WorkbookLoader()
        self.output_root.mkdir(parents=True, exist_ok=True)

    def build(self, session: ImportSession) -> Path:
        started_at = perf_counter()
        confirmed_rules = [
            rule for rule in session.mapping_rules if rule.confirmed and rule.source_header
        ]
        output_dir = self.output_root / session.id
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / "itogovyj_bdr.xlsx"

        workbook = self.loader.create_output_workbook()
        bdr_sheet = workbook.active
        bdr_sheet.title = "БДР"
        report_sheet = workbook.create_sheet("Отчет")
        source_sheet = workbook.create_sheet("Исходные данные")

        bdr_rows = self._extract_bdr_rows(session, confirmed_rules)
        self._write_bdr_sheet(session, bdr_sheet, confirmed_rules, bdr_rows)
        self._validate_bdr_sheet(session, bdr_sheet, confirmed_rules)
        self._write_report_sheet(session, report_sheet)
        self._write_source_summary(session, source_sheet)
        self._format_workbook(workbook)

        session.output_path = self.loader.save_output_workbook(workbook, output_path)
        session.performance.append(
            PerformanceEntry(
                stage="bdr_build",
                duration_seconds=round(perf_counter() - started_at, 6),
                sheets_processed=len({rule.source_sheet for rule in confirmed_rules if rule.source_sheet}),
                rows_read=max(bdr_sheet.max_row - 1, 0),
                cells_read=max(bdr_sheet.max_row - 1, 0) * max(bdr_sheet.max_column, 0),
            )
        )
        return session.output_path

    def _write_bdr_sheet(
        self,
        session: ImportSession,
        sheet: Any,
        rules: list[MappingRule],
        bdr_rows: list[dict[str, Any]],
    ) -> None:
        if not rules:
            session.decisions.append(
                ImportDecision(
                    severity=Severity.WARNING,
                    code="маппинг_не_задан",
                    message="Итоговый БДР создан без данных: нет выбранных правил сопоставления.",
                )
            )
            sheet.append(["Нет выбранных правил сопоставления"])
            return

        target_fields = [rule.target_field for rule in rules]
        sheet.append(target_fields)
        neutralized_values = 0
        for source_row in bdr_rows:
            output_row = []
            for target_field in target_fields:
                value = source_row.get(target_field)
                safe_value = neutralize_excel_formula(value)
                if safe_value != value:
                    neutralized_values += 1
                output_row.append(safe_value)
            sheet.append(output_row)

        if neutralized_values:
            session.decisions.append(
                ImportDecision(
                    severity=Severity.WARNING,
                    code="excel_формулы_нейтрализованы",
                    message=(
                        "В итоговом БДР найдены значения, похожие на Excel-формулы. "
                        "Они записаны как текст для безопасного открытия файла."
                    ),
                    context={"значений": neutralized_values},
                )
            )

        if not bdr_rows:
            session.decisions.append(
                ImportDecision(
                    severity=Severity.WARNING,
                    code="строки_источника_не_найдены",
                    message="Правила сопоставления выбраны, но строки данных не найдены.",
                )
            )

    def _extract_bdr_rows(
        self,
        session: ImportSession,
        rules: list[MappingRule],
    ) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        build_limit_reached = False
        missing_formula_cache_count = 0
        profiles_by_id = {profile.file_id: profile for profile in session.profiles}
        rules_by_source: dict[tuple[str, str], list[MappingRule]] = defaultdict(list)
        for rule in rules:
            if rule.source_file_id and rule.source_sheet and rule.source_header:
                rules_by_source[(rule.source_file_id, rule.source_sheet)].append(rule)

        for (file_id, sheet_name), source_rules in rules_by_source.items():
            profile = profiles_by_id.get(file_id)
            if profile is None:
                continue
            sheet_profile = next((sheet for sheet in profile.sheets if sheet.name == sheet_name), None)
            if sheet_profile is None or sheet_profile.header_row is None:
                for rule in source_rules:
                    session.decisions.append(
                        ImportDecision(
                            severity=Severity.WARNING,
                            code="лист_для_правила_не_найден",
                            message=f"Не удалось прочитать источник для поля «{rule.target_field}».",
                            context=rule.model_dump(mode="json"),
                        )
                    )
                continue

            with (
                self.loader.open_for_build(profile.path) as workbook,
                self.loader.open_for_build_formulas(profile.path) as formula_workbook,
            ):
                worksheet = workbook[sheet_name]
                formula_worksheet = formula_workbook[sheet_name]
                header_values = next(
                    worksheet.iter_rows(
                        min_row=sheet_profile.header_row,
                        max_row=sheet_profile.header_row,
                        values_only=True,
                    ),
                    (),
                )
                headers_by_column = {
                    str(value or "").strip(): column_index
                    for column_index, value in enumerate(header_values, start=1)
                    if str(value or "").strip()
                }
                column_by_target: dict[str, int] = {}
                missing_rules: list[MappingRule] = []
                for rule in source_rules:
                    column_index = headers_by_column.get(rule.source_header or "")
                    if column_index is None:
                        missing_rules.append(rule)
                    else:
                        column_by_target[rule.target_field] = column_index

                for rule in missing_rules:
                    session.decisions.append(
                        ImportDecision(
                            severity=Severity.WARNING,
                            code="колонка_для_правила_не_найдена",
                            message=f"Колонка «{rule.source_header}» не найдена при сборке поля «{rule.target_field}».",
                            context=rule.model_dump(mode="json"),
                        )
                    )
                if not column_by_target:
                    continue

                min_column = min(column_by_target.values())
                max_column = max(column_by_target.values())
                relative_columns = {
                    target_field: column_index - min_column
                    for target_field, column_index in column_by_target.items()
                }
                value_rows = worksheet.iter_rows(
                    min_row=sheet_profile.header_row + 1,
                    max_row=worksheet.max_row,
                    min_col=min_column,
                    max_col=max_column,
                    values_only=True,
                )
                formula_rows = formula_worksheet.iter_rows(
                    min_row=sheet_profile.header_row + 1,
                    max_row=worksheet.max_row,
                    min_col=min_column,
                    max_col=max_column,
                    values_only=True,
                )
                for values, formula_values in zip(value_rows, formula_rows, strict=False):
                    source_row = {
                        target_field: self._resolve_build_value(
                            values[relative_index],
                            formula_values[relative_index],
                        )
                        for target_field, relative_index in relative_columns.items()
                    }
                    missing_formula_cache_count += sum(
                        1
                        for relative_index in relative_columns.values()
                        if self._formula_cache_missing(
                            values[relative_index],
                            formula_values[relative_index],
                        )
                    )
                    if any(value not in (None, "") for value in source_row.values()):
                        rows.append(source_row)
                        if len(rows) >= MAX_BUILD_ROWS:
                            build_limit_reached = True
                            break
                if build_limit_reached:
                    break

            if build_limit_reached:
                session.decisions.append(
                    ImportDecision(
                        severity=Severity.WARNING,
                        code="лимит_строк_сборки",
                        message=(
                            "Сборка БДР остановлена по лимиту строк. "
                            f"В итоговый файл включено первых {MAX_BUILD_ROWS} строк."
                        ),
                        context={"лимит": MAX_BUILD_ROWS},
                    )
                )
                break
        if missing_formula_cache_count:
            session.decisions.append(
                ImportDecision(
                    severity=Severity.WARNING,
                    code="формулы_без_сохраненных_значений",
                    message=(
                        "В исходных файлах найдены формулы без сохраненных расчетных значений. "
                        "В итоговый БДР они добавлены как текст, чтобы не потерять данные молча."
                    ),
                    context={"ячеек": missing_formula_cache_count},
                )
            )
        return rows

    def _resolve_build_value(self, cached_value: Any, formula_value: Any) -> Any:
        if self._formula_cache_missing(cached_value, formula_value):
            return formula_value
        return cached_value

    def _formula_cache_missing(self, cached_value: Any, formula_value: Any) -> bool:
        return cached_value is None and isinstance(formula_value, str) and formula_value.startswith("=")

    def _validate_bdr_sheet(
        self,
        session: ImportSession,
        sheet: Any,
        rules: list[MappingRule],
    ) -> None:
        if pa is None or not rules or sheet.max_row <= 1:
            return

        headers = [cell.value for cell in sheet[1]]
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        dataframe = pd.DataFrame(rows, columns=headers)
        checks: dict[str, Any] = {}
        if "Статья БДР" in dataframe.columns:
            checks["Статья БДР"] = pa.Column(str, nullable=False)
        if "Сумма" in dataframe.columns:
            original_amount = dataframe["Сумма"]
            dataframe["Сумма"] = pd.to_numeric(original_amount, errors="coerce")
            invalid_amount_mask = original_amount.notna() & dataframe["Сумма"].isna()
            if invalid_amount_mask.any():
                session.decisions.append(
                    ImportDecision(
                        severity=Severity.WARNING,
                        code="некорректные_суммы",
                        message="В колонке «Сумма» найдены значения, которые не удалось преобразовать в число.",
                        context={"rows": invalid_amount_mask[invalid_amount_mask].index.tolist()},
                    )
                )
            checks["Сумма"] = pa.Column(float, nullable=True)
        if "Период" in dataframe.columns:
            checks["Период"] = pa.Column(object, nullable=True)
        if not checks:
            return

        started_at = perf_counter()
        schema = pa.DataFrameSchema(checks, coerce=True)
        try:
            schema.validate(dataframe, lazy=True)
        except pa.errors.SchemaErrors as exc:
            session.decisions.append(
                ImportDecision(
                    severity=Severity.WARNING,
                    code="валидация_dataframe_с_замечаниями",
                    message="Pandera-валидация итогового БДР нашла спорные значения.",
                    context={"failure_count": len(exc.failure_cases)},
                )
            )
        finally:
            session.performance.append(
                PerformanceEntry(
                    stage="schema_validation",
                    duration_seconds=round(perf_counter() - started_at, 6),
                    rows_read=len(dataframe),
                    cells_read=int(dataframe.size),
                )
            )

    def _write_report_sheet(self, session: ImportSession, sheet: Any) -> None:
        sheet.append(["Тип", "Код", "Сообщение"])
        for decision in session.decisions:
            sheet.append([decision.severity.value, decision.code, decision.message])
        if len(session.decisions) == 0:
            sheet.append(["информация", "замечаний_нет", "Ошибки и предупреждения не найдены."])

    def _write_source_summary(self, session: ImportSession, sheet: Any) -> None:
        sheet.append(["Файл", "Роль", "Уверенность", "Листов", "Формул", "Строки решений"])
        for profile in session.profiles:
            sheet.append(
                [
                    profile.original_name,
                    profile.detected_role.value,
                    profile.role_confidence,
                    len(profile.sheets),
                    profile.fingerprint.total_formula_cells,
                    "; ".join(profile.role_reasons),
                ]
            )
        sheet.append([])
        sheet.append(["Целевое поле", "Файл", "Лист", "Колонка", "Подтверждено"])
        profiles_by_id = {profile.file_id: profile for profile in session.profiles}
        for rule in session.mapping_rules:
            profile = profiles_by_id.get(rule.source_file_id or "")
            sheet.append(
                [
                    rule.target_field,
                    profile.original_name if profile else "",
                    rule.source_sheet or "",
                    rule.source_header or "",
                    "да" if rule.confirmed else "нет",
                ]
            )

    def _format_workbook(self, workbook: Workbook) -> None:
        header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
        header_font = Font(bold=True, color="1F2933")
        for sheet in workbook.worksheets:
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            widths: dict[int, int] = defaultdict(lambda: 10)
            for row in sheet.iter_rows():
                for cell in row:
                    value = str(cell.value or "")
                    widths[cell.column] = min(48, max(widths[cell.column], len(value) + 2))
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
            for column_index, width in widths.items():
                sheet.column_dimensions[get_column_letter(column_index)].width = width
            sheet.freeze_panes = "A2"
