from __future__ import annotations

import hashlib
import re
from collections import Counter, defaultdict
from pathlib import Path
from statistics import mean
from time import perf_counter

from openpyxl.worksheet.worksheet import Worksheet
from pydantic import TypeAdapter

from bdr_wizard.excel.loader import WorkbookLoader
from bdr_wizard.localization import TARGET_FIELDS
from bdr_wizard.models import (
    FileRole,
    FormulaSignature,
    ImportDecision,
    PerformanceEntry,
    Severity,
    SheetMeta,
    SheetProfile,
    UploadedFile,
    WorkbookFingerprint,
    WorkbookMeta,
    WorkbookProfile,
)
from bdr_wizard.security import MAX_WORKBOOK_SHEETS


CELL_REF_RE = re.compile(r"(?<![A-Za-zА-Яа-я0-9_])\$?[A-Z]{1,3}\$?\d+")
NUMBER_RE = re.compile(r"\b\d+(?:[.,]\d+)?\b")
STRING_RE = re.compile(r'"[^"]*"')


def normalize_text(value: object) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("ё", "е")
    text = re.sub(r"\s+", " ", text)
    return text


def normalize_formula(value: str) -> str:
    formula = value.upper().strip()
    formula = STRING_RE.sub('"TEXT"', formula)
    formula = CELL_REF_RE.sub("CELL", formula)
    formula = NUMBER_RE.sub("NUM", formula)
    return formula


class WorkbookAnalysisCache:
    def __init__(self, cache_root: Path = Path("data/cache/workbooks")) -> None:
        self.cache_root = cache_root
        self.cache_root.mkdir(parents=True, exist_ok=True)

    def file_key(self, path: Path) -> str:
        digest = hashlib.sha256()
        with path.open("rb") as handle:
            for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                digest.update(chunk)
        return digest.hexdigest()

    def get(self, key: str) -> WorkbookProfile | None:
        cache_path = self.cache_root / f"{key}.json"
        if not cache_path.exists():
            return None
        return TypeAdapter(WorkbookProfile).validate_json(cache_path.read_text("utf-8"))

    def set(self, key: str, profile: WorkbookProfile) -> None:
        cache_path = self.cache_root / f"{key}.json"
        temporary_path = cache_path.with_suffix(f".{profile.file_id}.tmp")
        temporary_path.write_text(profile.model_dump_json(), encoding="utf-8")
        temporary_path.replace(cache_path)


class WorkbookAnalysisError(RuntimeError):
    def __init__(self, file_name: str, message: str) -> None:
        self.file_name = file_name
        super().__init__(f"Не удалось проанализировать файл «{file_name}»: {message}")


class WorkbookAnalyzer:
    def __init__(
        self,
        cache: WorkbookAnalysisCache | None = None,
        loader: WorkbookLoader | None = None,
        header_scan_rows: int = 25,
        structural_scan_rows: int = 200,
        max_scan_columns: int = 80,
    ) -> None:
        self.cache = cache or WorkbookAnalysisCache()
        self.loader = loader or WorkbookLoader()
        self.header_scan_rows = header_scan_rows
        self.structural_scan_rows = structural_scan_rows
        self.max_scan_columns = max_scan_columns

    def analyze(self, uploaded_file: UploadedFile) -> WorkbookProfile:
        started_at = perf_counter()
        cache_key = self.cache.file_key(uploaded_file.stored_path)
        cached_profile = self.cache.get(cache_key)
        if cached_profile is not None:
            profile = cached_profile.model_copy(
                update={
                    "file_id": uploaded_file.id,
                    "original_name": uploaded_file.original_name,
                    "path": uploaded_file.stored_path,
                },
                deep=True,
            )
            profile.performance = [
                PerformanceEntry(
                    stage="structural_screening",
                    duration_seconds=round(perf_counter() - started_at, 6),
                    file_id=uploaded_file.id,
                    file_name=uploaded_file.original_name,
                    sheets_processed=len(profile.sheets),
                    cache_status="hit",
                )
            ]
            return profile

        rows_read = 0
        cells_read = 0
        try:
            with self.loader.open_for_analysis(uploaded_file.stored_path) as workbook:
                if len(workbook.worksheets) > MAX_WORKBOOK_SHEETS:
                    raise WorkbookAnalysisError(
                        uploaded_file.original_name,
                        "в книге слишком много листов для безопасной автоматической обработки.",
                    )
                sheets: list[SheetProfile] = []
                for sheet in workbook.worksheets:
                    profile, sheet_rows, sheet_cells = self._profile_sheet(sheet)
                    sheets.append(profile)
                    rows_read += sheet_rows
                    cells_read += sheet_cells

            fingerprint = self._build_fingerprint(sheets)
            detected_role, confidence, reasons = self._detect_role(sheets, fingerprint)
            decisions = self._build_decisions(sheets, confidence)
            profile = WorkbookProfile(
                file_id=uploaded_file.id,
                original_name=uploaded_file.original_name,
                path=uploaded_file.stored_path,
                sheets=sheets,
                fingerprint=fingerprint,
                detected_role=detected_role,
                role_confidence=confidence,
                role_reasons=reasons,
                decisions=decisions,
                performance=[
                    PerformanceEntry(
                        stage="structural_screening",
                        duration_seconds=round(perf_counter() - started_at, 6),
                        file_id=uploaded_file.id,
                        file_name=uploaded_file.original_name,
                        sheets_processed=len(sheets),
                        rows_read=rows_read,
                        cells_read=cells_read,
                        cache_status="miss",
                    )
                ],
            )
            profile_meta = build_workbook_meta(profile)
            profile.performance[0].details["workbook_meta_sheets"] = len(profile_meta.sheets)
            self.cache.set(cache_key, profile)
            return profile
        except WorkbookAnalysisError:
            raise
        except Exception as exc:
            raise WorkbookAnalysisError(
                uploaded_file.original_name,
                "проверьте, что файл является корректной Excel-книгой .xlsx или .xlsm.",
            ) from exc

    def _profile_sheet(self, sheet: Worksheet) -> tuple[SheetProfile, int, int]:
        non_empty_cells = 0
        formula_cells = 0
        formula_patterns: Counter[str] = Counter()
        row_signatures: Counter[str] = Counter()
        rows_read = 0
        cells_read = 0

        max_row = sheet.max_row or 0
        max_column = sheet.max_column or 0
        scan_rows = min(max_row, self.structural_scan_rows)
        scan_columns = min(max_column, self.max_scan_columns)
        scan_truncated = max_row > scan_rows or max_column > scan_columns

        if scan_rows and scan_columns:
            for row in sheet.iter_rows(
                min_row=1,
                max_row=scan_rows,
                min_col=1,
                max_col=scan_columns,
            ):
                rows_read += 1
                signature_parts: list[str] = []
                for cell in row:
                    cells_read += 1
                    if cell.value not in (None, ""):
                        non_empty_cells += 1
                        if isinstance(cell.value, str) and cell.value.startswith("="):
                            formula_cells += 1
                            pattern = normalize_formula(cell.value)
                            formula_patterns[pattern] += 1
                            signature_parts.append("F")
                        else:
                            signature_parts.append("V")
                    else:
                        signature_parts.append("_")
                compact_signature = "".join(signature_parts).rstrip("_")
                if compact_signature:
                    row_signatures[compact_signature] += 1

        header_row, headers, header_rows_read, header_cells_read = self._detect_headers(
            sheet,
            max_column=scan_columns,
        )
        rows_read += header_rows_read
        cells_read += header_cells_read

        repeated_blocks = [
            f"Строки с похожей структурой: {signature} ({count} раз)"
            for signature, count in row_signatures.most_common(5)
            if count >= 3 and len(signature) >= 2
        ]
        capacity = max(scan_rows * max(scan_columns, 1), 1)
        hidden = getattr(sheet, "sheet_state", "visible") != "visible"
        empty = max_row <= 1 and max_column <= 1 and non_empty_cells == 0
        candidate_for_processing = bool(headers) and not hidden and not empty

        return (
            SheetProfile(
                name=sheet.title,
                max_row=max_row,
                max_column=max_column,
                non_empty_cells=non_empty_cells,
                formula_cells=formula_cells,
                data_density=round(non_empty_cells / capacity, 4),
                scanned_rows=rows_read,
                scanned_cells=cells_read,
                scan_truncated=scan_truncated,
                hidden=hidden,
                empty=empty,
                candidate_for_processing=candidate_for_processing,
                header_row=header_row,
                headers=headers,
                formula_patterns=[pattern for pattern, _ in formula_patterns.most_common(15)],
                table_ranges=[],
                repeated_blocks=repeated_blocks,
            ),
            rows_read,
            cells_read,
        )

    def _detect_headers(
        self,
        sheet: Worksheet,
        max_column: int,
    ) -> tuple[int | None, list[str], int, int]:
        best_row: int | None = None
        best_headers: list[str] = []
        best_score = 0.0
        scan_limit = min(sheet.max_row or 0, self.header_scan_rows)
        rows_read = 0
        cells_read = 0

        if not scan_limit or not max_column:
            return None, [], rows_read, cells_read

        for row_index, row in enumerate(
            sheet.iter_rows(
                min_row=1,
                max_row=scan_limit,
                min_col=1,
                max_col=max_column,
            ),
            start=1,
        ):
            rows_read += 1
            values = [cell.value for cell in row]
            cells_read += len(values)
            texts = [str(value).strip() for value in values if isinstance(value, str) and value.strip()]
            if not texts:
                continue
            filled = len([value for value in values if value not in (None, "")])
            text_ratio = len(texts) / max(filled, 1)
            uniqueness = len(set(normalize_text(text) for text in texts)) / max(len(texts), 1)
            domain_hits = sum(self._looks_like_domain_header(text) for text in texts)
            score = filled + text_ratio * 4 + uniqueness * 2 + domain_hits * 3
            if score > best_score:
                best_score = score
                best_row = row_index
                best_headers = texts

        if best_row is None or best_score < 3:
            return None, [], rows_read, cells_read
        return best_row, best_headers, rows_read, cells_read

    def _looks_like_domain_header(self, text: str) -> bool:
        normalized = normalize_text(text)
        for synonyms in TARGET_FIELDS.values():
            if any(synonym in normalized for synonym in synonyms):
                return True
        return False

    def _build_fingerprint(self, sheets: list[SheetProfile]) -> WorkbookFingerprint:
        header_tokens = sorted(
            {
                normalize_text(header)
                for sheet in sheets
                for header in sheet.headers
                if normalize_text(header)
            }
        )
        formula_tokens = sorted(
            {
                pattern
                for sheet in sheets
                for pattern in sheet.formula_patterns
            }
        )
        structure_parts = [
            (
                f"{sheet.name}:{sheet.max_row}x{sheet.max_column}:"
                f"{sheet.hidden}:{sheet.empty}:{sheet.non_empty_cells}:{sheet.formula_cells}"
            )
            for sheet in sheets
        ]
        signature_source = "|".join(structure_parts + header_tokens[:50] + formula_tokens[:50])
        structure_signature = hashlib.sha256(signature_source.encode("utf-8")).hexdigest()[:16]
        return WorkbookFingerprint(
            sheet_count=len(sheets),
            total_non_empty_cells=sum(sheet.non_empty_cells for sheet in sheets),
            total_formula_cells=sum(sheet.formula_cells for sheet in sheets),
            header_tokens=header_tokens,
            formula_tokens=formula_tokens,
            structure_signature=structure_signature,
        )

    def _detect_role(
        self,
        sheets: list[SheetProfile],
        fingerprint: WorkbookFingerprint,
    ) -> tuple[FileRole, float, list[str]]:
        scores: dict[FileRole, float] = defaultdict(float)
        reasons: dict[FileRole, list[str]] = defaultdict(list)
        candidate_sheets = [sheet for sheet in sheets if sheet.candidate_for_processing]
        headers = " ".join(fingerprint.header_tokens)
        avg_density = mean([sheet.data_density for sheet in candidate_sheets]) if candidate_sheets else 0

        if fingerprint.total_formula_cells > 0:
            scores[FileRole.MAIN] += 2.0
            reasons[FileRole.MAIN].append("найдены формулы и расчетные структуры")
        if any(token in headers for token in ["статья", "период", "сумма", "план", "факт", "бюджет"]):
            scores[FileRole.MAIN] += 2.0
            reasons[FileRole.MAIN].append("найдены признаки таблицы БДР")
        if any(token in headers for token in ["источник", "целев", "маппинг", "соответствие"]):
            scores[FileRole.MAPPING] += 3.0
            reasons[FileRole.MAPPING].append("найдены признаки таблицы сопоставлений")
        if any(token in headers for token in ["код", "наименование", "справочник", "цфо"]):
            scores[FileRole.DIRECTORY] += 2.0
            reasons[FileRole.DIRECTORY].append("найдены признаки справочника")
        if fingerprint.total_formula_cells > 5 and fingerprint.sheet_count >= 2:
            scores[FileRole.TEMPLATE] += 2.0
            reasons[FileRole.TEMPLATE].append("много формул и несколько листов")
        if avg_density < 0.08 and fingerprint.total_non_empty_cells > 0:
            scores[FileRole.TEMPLATE] += 1.0
            reasons[FileRole.TEMPLATE].append("разреженная структура похожа на шаблон")

        if not scores:
            return FileRole.OTHER, 0.25, ["недостаточно признаков для уверенного определения"]

        role = max(scores, key=scores.get)
        sorted_scores = sorted(scores.values(), reverse=True)
        top_score = sorted_scores[0]
        second_score = sorted_scores[1] if len(sorted_scores) > 1 else 0
        confidence = min(0.95, max(0.35, (top_score - second_score + 1) / (top_score + 1)))
        return role, round(confidence, 2), reasons[role]

    def _build_decisions(
        self,
        sheets: list[SheetProfile],
        confidence: float,
    ) -> list[ImportDecision]:
        decisions: list[ImportDecision] = []
        if confidence < 0.6:
            decisions.append(
                ImportDecision(
                    severity=Severity.WARNING,
                    code="неуверенная_роль",
                    message="Роль файла определена неуверенно. Проверьте выбор на экране предпросмотра.",
                )
            )
        for sheet in sheets:
            if sheet.hidden:
                decisions.append(
                    ImportDecision(
                        severity=Severity.INFO,
                        code="скрытый_лист_пропущен",
                        message=f"Лист «{sheet.name}» скрыт и не включен в углубленную обработку.",
                    )
                )
            if sheet.empty:
                decisions.append(
                    ImportDecision(
                        severity=Severity.INFO,
                        code="пустой_лист_пропущен",
                        message=f"Лист «{sheet.name}» выглядит пустым и не включен в углубленную обработку.",
                    )
                )
            if not sheet.headers and not sheet.empty:
                decisions.append(
                    ImportDecision(
                        severity=Severity.WARNING,
                        code="заголовки_не_найдены",
                        message=f"На листе «{sheet.name}» не удалось надежно определить строку заголовков.",
                    )
                )
            if sheet.scan_truncated:
                decisions.append(
                    ImportDecision(
                        severity=Severity.INFO,
                        code="лист_просканирован_частично",
                        message=(
                            f"Лист «{sheet.name}» просканирован в быстром режиме частично: "
                            f"{sheet.scanned_rows} строк, {sheet.scanned_cells} ячеек."
                        ),
                    )
                )
        return decisions


def build_workbook_meta(profile: WorkbookProfile) -> WorkbookMeta:
    return WorkbookMeta(
        file_id=profile.file_id,
        original_name=profile.original_name,
        path=profile.path,
        fingerprint=profile.fingerprint,
        sheets=[
            SheetMeta(
                name=sheet.name,
                max_row=sheet.max_row,
                max_column=sheet.max_column,
                hidden=sheet.hidden,
                empty=sheet.empty,
                headers=sheet.headers,
                formula_signatures=[
                    FormulaSignature(sheet_name=sheet.name, pattern=pattern)
                    for pattern in sheet.formula_patterns
                ],
            )
            for sheet in profile.sheets
        ],
    )
