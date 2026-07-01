from __future__ import annotations

from collections.abc import Iterator
from contextlib import contextmanager
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.workbook.workbook import Workbook as OpenpyxlWorkbook

from bdr_wizard.security import validate_excel_archive
from bdr_wizard.storage.encryption import EncryptedFileStore


class WorkbookLoader:
    """Единая точка открытия и создания Excel workbook через openpyxl."""

    def __init__(self, encrypted_store: EncryptedFileStore | None = None) -> None:
        self.encrypted_store = encrypted_store or EncryptedFileStore()

    @contextmanager
    def open_for_analysis(self, path: Path) -> Iterator[OpenpyxlWorkbook]:
        with self._open(path, data_only=False) as workbook:
            yield workbook

    @contextmanager
    def open_for_build(self, path: Path) -> Iterator[OpenpyxlWorkbook]:
        with self._open(path, data_only=True) as workbook:
            yield workbook

    @contextmanager
    def open_for_build_formulas(self, path: Path) -> Iterator[OpenpyxlWorkbook]:
        with self._open(path, data_only=False) as workbook:
            yield workbook

    @contextmanager
    def _open(self, path: Path, data_only: bool) -> Iterator[OpenpyxlWorkbook]:
        suffix = "".join(path.suffixes).removesuffix(".enc")
        with self.encrypted_store.materialized_file(path, suffix=suffix) as readable_path:
            validate_excel_archive(readable_path)
            workbook = load_workbook(
                readable_path,
                data_only=data_only,
                read_only=True,
                keep_links=False,
            )
            try:
                yield workbook
            finally:
                workbook.close()

    def save_output_workbook(self, workbook: Workbook, path: Path) -> Path:
        temporary_path = path.with_suffix(f"{path.suffix}.tmp")
        try:
            workbook.save(temporary_path)
            return self.encrypted_store.write_encrypted(path, temporary_path.read_bytes())
        finally:
            temporary_path.unlink(missing_ok=True)

    def create_output_workbook(self) -> Workbook:
        return Workbook()
