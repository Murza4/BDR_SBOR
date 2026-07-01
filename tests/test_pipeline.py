from __future__ import annotations

from pathlib import Path
import json

from openpyxl import Workbook, load_workbook

from bdr_wizard.auth import AuthSettings, create_auth_token, verify_auth_token, verify_credentials
from bdr_wizard.analyzer import WorkbookAnalysisError, WorkbookAnalyzer
from bdr_wizard.analyzer.service import WorkbookAnalysisCache
from bdr_wizard.builder import BdrBuilder
from bdr_wizard.excel.compare import CompareEngine
from bdr_wizard.excel.loader import WorkbookLoader
from bdr_wizard.ingestion import IngestionService
from bdr_wizard.mapper import MappingEngine, MappingRuleStore
from bdr_wizard.models import FileRole
from bdr_wizard.reporting import ReportingService
from bdr_wizard.security import MAX_UPLOAD_SIZE_BYTES, validate_upload_file_count
from bdr_wizard.storage.encryption import EncryptedFileStore


def _make_bdr_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "План"
    sheet.append(["Статья БДР", "Код статьи", "Период", "Сумма", "Подразделение"])
    sheet.append(["Выручка", "100", "2026-01", 1200, "Продажи"])
    sheet.append(["Расходы", "200", "2026-01", 350, "Офис"])
    sheet["F1"] = "Расчет"
    sheet["F2"] = "=D2*1.2"
    sheet["F3"] = "=D3*1.2"
    workbook.save(path)


def _make_bdr_workbook_with_blank_amount(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "План"
    sheet.append(["Статья БДР", "Период", "Сумма"])
    sheet.append(["Выручка", "2026-01", 1200])
    sheet.append(["Расходы", "2026-01", None])
    sheet.append(["Маржа", "2026-01", 850])
    workbook.save(path)


def _make_bdr_workbook_with_dangerous_text(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "План"
    sheet.append(["Статья БДР", "Период", "Сумма"])
    sheet.append(["+Небезопасный текст", "2026-01", 100])
    workbook.save(path)


def _make_bdr_workbook_with_formula_amount(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "План"
    sheet.append(["Статья БДР", "Период", "Сумма"])
    sheet.append(["Выручка", "2026-01", "=100+200"])
    workbook.save(path)


def _make_simple_workbook(path: Path, headers: list[str]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Данные"
    sheet.append(headers)
    sheet.append(["A", "B", "C"])
    workbook.save(path)


def _load_workbook(path: Path, data_only: bool = True):
    store = EncryptedFileStore()
    with store.materialized_file(path, suffix=".xlsx") as readable_path:
        workbook = load_workbook(readable_path, data_only=data_only)
    return workbook


def _read_text(path: Path) -> str:
    return EncryptedFileStore().read_bytes(path).decode("utf-8")


def test_analyzer_detects_headers_formulas_and_role(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    _make_bdr_workbook(source)
    ingestion = IngestionService(tmp_path / "uploads")
    session = ingestion.create_session()
    uploaded = ingestion.add_file(session, source)

    profile = WorkbookAnalyzer().analyze(uploaded)

    assert profile.detected_role in {FileRole.MAIN, FileRole.TEMPLATE}
    assert profile.fingerprint.total_formula_cells == 2
    assert profile.sheets[0].header_row == 1
    assert "Статья БДР" in profile.sheets[0].headers
    assert profile.performance[0].cache_status == "miss"
    assert profile.performance[0].cells_read > 0
    assert profile.sheets[0].candidate_for_processing is True
    assert profile.sheets[0].scanned_cells > 0


def test_analyzer_reuses_cached_fingerprint(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    _make_bdr_workbook(source)
    ingestion = IngestionService(tmp_path / "uploads")
    session = ingestion.create_session()
    uploaded = ingestion.add_file(session, source)
    analyzer = WorkbookAnalyzer(WorkbookAnalysisCache(tmp_path / "cache"))

    first_profile = analyzer.analyze(uploaded)
    second_profile = analyzer.analyze(uploaded)

    assert first_profile.fingerprint.structure_signature == second_profile.fingerprint.structure_signature
    assert first_profile.performance[0].cache_status == "miss"
    assert second_profile.performance[0].cache_status == "hit"


def test_workbook_loader_closes_analysis_workbook(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    _make_bdr_workbook(source)
    loader = WorkbookLoader()

    with loader.open_for_analysis(source) as workbook:
        assert workbook.read_only is True
        assert workbook.sheetnames == ["План"]

    assert workbook._archive.fp is None


def test_compare_engine_reports_similar_workbooks(tmp_path: Path) -> None:
    first_source = tmp_path / "first.xlsx"
    second_source = tmp_path / "second.xlsx"
    _make_bdr_workbook(first_source)
    _make_bdr_workbook(second_source)
    ingestion = IngestionService(tmp_path / "uploads")
    session = ingestion.create_session()
    first = ingestion.add_file(session, first_source)
    second = ingestion.add_file(session, second_source)
    analyzer = WorkbookAnalyzer(WorkbookAnalysisCache(tmp_path / "cache"))

    report = CompareEngine().compare([analyzer.analyze(first), analyzer.analyze(second)])

    assert report.similar_pairs
    assert report.decisions[0].code == "похожие_файлы"


def test_mapping_engine_suggests_core_fields(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    _make_bdr_workbook(source)
    ingestion = IngestionService(tmp_path / "uploads")
    session = ingestion.create_session()
    uploaded = ingestion.add_file(session, source)
    profile = WorkbookAnalyzer().analyze(uploaded)

    engine = MappingEngine(MappingRuleStore(tmp_path / "rules.json"))
    rules, candidates = engine.suggest([profile])

    assert candidates
    by_field = {rule.target_field: rule for rule in rules}
    assert by_field["Статья БДР"].source_header == "Статья БДР"
    assert by_field["Сумма"].source_header == "Сумма"


def test_mapping_rules_are_reused_with_new_file_ids(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    _make_bdr_workbook(source)
    ingestion = IngestionService(tmp_path / "uploads")
    analyzer = WorkbookAnalyzer(WorkbookAnalysisCache(tmp_path / "cache"))
    rule_store = MappingRuleStore(tmp_path / "rules.json")
    engine = MappingEngine(rule_store)

    first_session = ingestion.create_session()
    first_uploaded = ingestion.add_file(first_session, source)
    first_profile = analyzer.analyze(first_uploaded)
    first_rules, _ = engine.suggest([first_profile])
    confirmed_rules = [
        rule.model_copy(update={"confirmed": True})
        for rule in first_rules
        if rule.source_header
    ]
    rule_store.save(confirmed_rules)

    second_session = ingestion.create_session()
    second_uploaded = ingestion.add_file(second_session, source)
    second_profile = analyzer.analyze(second_uploaded)
    second_rules, _ = engine.suggest([second_profile])
    by_field = {rule.target_field: rule for rule in second_rules}

    assert second_uploaded.id != first_uploaded.id
    assert by_field["Статья БДР"].source_file_id == second_uploaded.id
    assert by_field["Статья БДР"].source_signature == confirmed_rules[0].source_signature


def test_builder_creates_result_and_report(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    _make_bdr_workbook(source)
    ingestion = IngestionService(tmp_path / "uploads")
    session = ingestion.create_session()
    uploaded = ingestion.add_file(session, source)
    profile = WorkbookAnalyzer().analyze(uploaded)
    session.profiles.append(profile)

    engine = MappingEngine(MappingRuleStore(tmp_path / "rules.json"))
    rules, _ = engine.suggest(session.profiles)
    session.mapping_rules = [
        rule.model_copy(update={"confirmed": True})
        for rule in rules
        if rule.source_header in {"Статья БДР", "Сумма", "Период"}
    ]

    output_path = BdrBuilder(tmp_path / "outputs").build(session)
    report_path = ReportingService(tmp_path / "reports").write_audit_report(session)

    assert output_path.exists()
    assert report_path.exists()
    report = json.loads(_read_text(report_path))
    assert "производительность" in report
    assert any(entry["stage"] == "bdr_build" for entry in report["производительность"])
    result = _load_workbook(output_path)
    assert "БДР" in result.sheetnames
    assert "Отчет" in result.sheetnames
    assert result["БДР"]["A1"].value == "Статья БДР"
    assert result["БДР"]["A2"].value == "Выручка"


def test_builder_preserves_row_alignment_with_blank_values(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    _make_bdr_workbook_with_blank_amount(source)
    ingestion = IngestionService(tmp_path / "uploads")
    session = ingestion.create_session()
    uploaded = ingestion.add_file(session, source)
    profile = WorkbookAnalyzer(WorkbookAnalysisCache(tmp_path / "cache")).analyze(uploaded)
    session.profiles.append(profile)

    engine = MappingEngine(MappingRuleStore(tmp_path / "rules.json"))
    rules, _ = engine.suggest(session.profiles)
    session.mapping_rules = [
        rule.model_copy(update={"confirmed": True})
        for rule in rules
        if rule.source_header in {"Статья БДР", "Сумма", "Период"}
    ]
    output_path = BdrBuilder(tmp_path / "outputs").build(session)

    result = _load_workbook(output_path)
    rows = list(result["БДР"].iter_rows(values_only=True))
    headers = rows[0]
    article_index = headers.index("Статья БДР")
    amount_index = headers.index("Сумма")
    assert rows[2][article_index] == "Расходы"
    assert rows[2][amount_index] is None
    assert rows[3][article_index] == "Маржа"
    assert rows[3][amount_index] == 850


def test_builder_uses_only_confirmed_rules(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    _make_bdr_workbook(source)
    ingestion = IngestionService(tmp_path / "uploads")
    session = ingestion.create_session()
    uploaded = ingestion.add_file(session, source)
    profile = WorkbookAnalyzer(WorkbookAnalysisCache(tmp_path / "cache")).analyze(uploaded)
    session.profiles.append(profile)

    engine = MappingEngine(MappingRuleStore(tmp_path / "rules.json"))
    rules, _ = engine.suggest(session.profiles)
    session.mapping_rules = [rule for rule in rules if rule.source_header]

    output_path = BdrBuilder(tmp_path / "outputs").build(session)

    result = _load_workbook(output_path)
    assert result["БДР"]["A1"].value == "Нет выбранных правил сопоставления"


def test_builder_neutralizes_formula_like_text(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    _make_bdr_workbook_with_dangerous_text(source)
    ingestion = IngestionService(tmp_path / "uploads")
    session = ingestion.create_session()
    uploaded = ingestion.add_file(session, source)
    profile = WorkbookAnalyzer(WorkbookAnalysisCache(tmp_path / "cache")).analyze(uploaded)
    session.profiles.append(profile)

    engine = MappingEngine(MappingRuleStore(tmp_path / "rules.json"))
    rules, _ = engine.suggest(session.profiles)
    session.mapping_rules = [
        rule.model_copy(update={"confirmed": True})
        for rule in rules
        if rule.source_header in {"Статья БДР", "Сумма", "Период"}
    ]
    output_path = BdrBuilder(tmp_path / "outputs").build(session)

    result = _load_workbook(output_path, data_only=False)
    assert result["БДР"]["A2"].value == "'+Небезопасный текст"
    assert any(decision.code == "excel_формулы_нейтрализованы" for decision in session.decisions)


def test_builder_reports_formula_without_cached_value(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    _make_bdr_workbook_with_formula_amount(source)
    ingestion = IngestionService(tmp_path / "uploads")
    session = ingestion.create_session()
    uploaded = ingestion.add_file(session, source)
    profile = WorkbookAnalyzer(WorkbookAnalysisCache(tmp_path / "cache")).analyze(uploaded)
    session.profiles.append(profile)

    engine = MappingEngine(MappingRuleStore(tmp_path / "rules.json"))
    rules, _ = engine.suggest(session.profiles)
    session.mapping_rules = [
        rule.model_copy(update={"confirmed": True})
        for rule in rules
        if rule.source_header in {"Статья БДР", "Сумма", "Период"}
    ]

    output_path = BdrBuilder(tmp_path / "outputs").build(session)

    result = _load_workbook(output_path, data_only=False)
    headers = [cell.value for cell in result["БДР"][1]]
    amount_index = headers.index("Сумма") + 1
    assert result["БДР"].cell(row=2, column=amount_index).value == "'=100+200"
    assert any(decision.code == "формулы_без_сохраненных_значений" for decision in session.decisions)


def test_analyzer_reports_invalid_excel_as_controlled_error(tmp_path: Path) -> None:
    source = tmp_path / "broken.xlsx"
    source.write_text("не настоящая книга Excel", encoding="utf-8")
    ingestion = IngestionService(tmp_path / "uploads")
    session = ingestion.create_session()
    uploaded = ingestion.add_file(session, source)

    try:
        WorkbookAnalyzer(WorkbookAnalysisCache(tmp_path / "cache")).analyze(uploaded)
    except WorkbookAnalysisError as exc:
        assert "Не удалось проанализировать файл" in str(exc)
    else:
        raise AssertionError("Битый Excel должен возвращать контролируемую ошибку анализа.")


def test_ingestion_rejects_unsupported_extension(tmp_path: Path) -> None:
    source = tmp_path / "bad.txt"
    source.write_text("не Excel", encoding="utf-8")
    ingestion = IngestionService(tmp_path / "uploads")
    session = ingestion.create_session()

    try:
        ingestion.add_file(session, source)
    except ValueError as exc:
        assert "Поддерживаются только файлы Excel" in str(exc)
    else:
        raise AssertionError("Файл с неподдерживаемым расширением должен быть отклонен.")


def test_ingestion_rejects_oversized_upload(tmp_path: Path) -> None:
    source = tmp_path / "huge.xlsx"
    source.write_bytes(b"0" * (MAX_UPLOAD_SIZE_BYTES + 1))
    ingestion = IngestionService(tmp_path / "uploads")
    session = ingestion.create_session()

    try:
        ingestion.add_file(session, source)
    except ValueError as exc:
        assert "Файл слишком большой" in str(exc)
    else:
        raise AssertionError("Слишком большой файл должен быть отклонен до анализа.")


def test_ingestion_rejects_oversized_session(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr("bdr_wizard.security.MAX_SESSION_UPLOAD_SIZE_BYTES", 20)
    ingestion = IngestionService(tmp_path / "uploads")
    session = ingestion.create_session()
    content = b"0" * 10

    for index in range(2):
        ingestion.add_uploaded_bytes(session, f"file-{index}.xlsx", content)

    try:
        ingestion.add_uploaded_bytes(session, "too-much.xlsx", content)
    except ValueError as exc:
        assert "Суммарный размер файлов в сессии слишком большой" in str(exc)
    else:
        raise AssertionError("Сессия больше лимита должна быть отклонена.")


def test_upload_file_count_allows_up_to_seven_files() -> None:
    assert validate_upload_file_count(1) is None
    assert validate_upload_file_count(7) is None
    assert validate_upload_file_count(0) == "Выберите хотя бы один Excel-файл."
    assert validate_upload_file_count(8) == "Можно загрузить не больше 7 Excel-файлов за одну сессию."


def test_encrypted_file_store_round_trip(tmp_path: Path) -> None:
    store = EncryptedFileStore(tmp_path / "key")
    encrypted_path = store.write_encrypted(tmp_path / "source.xlsx", b"secret")

    assert encrypted_path.suffix == ".enc"
    assert encrypted_path.read_bytes() != b"secret"
    assert store.read_bytes(encrypted_path) == b"secret"


def test_auth_token_round_trip() -> None:
    settings = AuthSettings(disabled=False, username="admin", password="pass", secret="secret")
    token = create_auth_token("admin", settings)

    assert verify_credentials("admin", "pass", settings)
    assert not verify_credentials("admin", "bad", settings)
    assert verify_auth_token(token, settings)
    assert not verify_auth_token(token, AuthSettings(False, "admin", "pass", "other"))
